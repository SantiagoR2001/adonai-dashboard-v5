import os
import json
import pandas as pd
from datetime import datetime, date
from collections import defaultdict
import io
import openpyxl
from flask import Blueprint, render_template, request, jsonify, make_response, session, redirect, url_for, current_app

from .config import Config
from .excel_db import get_workbook, save_workbook, parse_date_cell, row_to_dict, actualizar_reportes

main_bp = Blueprint('main', __name__)

CAT_EN = {
    "1. Nómina": "1. Payroll",
    "2. Colaboradoras Internas": "2. Internal Collaborators",
    "3. Proveedores": "3. Suppliers",
    "4. Operacionales": "4. Operational",
    "5. Redes Sociales": "5. Social Media",
    "6. Material Publicitario": "6. Advertising Material",
    "7. Trabajo de Campo": "7. Field Work",
    "8. Otros Gastos": "8. Other Expenses",
    "9. Misional Bebés": "9. Missional Babies",
    "10. Misional Madres": "10. Missional Mothers",
    "Administrativo": "Administrative",
    "Misión Institucional": "Institutional Mission",
    "Unidad de Negocio": "Business Unit",
    "Otros ingresos": "Other Income",
    "Otros": "Other",
    # Subcategorías Ingresos
    "Donaciones SRL": "SRL Donations",
    "Inversiones": "Investments",
    "Donaciones locales": "Local Donations",
    "Donaciones específicas locales": "Specific Local Donations",
    "Donaciones USA": "USA Donations",
    "Donaciones específicas USA": "Specific USA Donations",
    "Donaciones ropa": "Clothing Donations",
    "Jabón": "Soap",
    "Intereses bancarios": "Bank Interest",
    # Subcategorías Egresos (Misional Madres/Bebes)
    "Kit para el parto": "Delivery Kit",
    "Medicamentos": "Medicines",
    "Alimentos": "Food",
    "Elementos básicos": "Basic Elements",
    "Cumpleaños": "Birthdays",
    "Jardín de la vida": "Garden of Life",
    "Kit para madres": "Mothers Kit",
    "Transporte": "Transportation",
    "Habitabilidad": "Housing",
    "Ayuda Humanitaria": "Humanitarian Aid",
    "Emprendimiento": "Entrepreneurship",
    "Obsequios": "Gifts",
    "Eventos": "Events",
    "Alianzas": "Alliances",
    "Dirección General": "General Direction",
    "Secretaria": "Secretary",
    "Seguridad Social": "Social Security",
    "Prestaciones Sociales": "Social Benefits",
    "Dotación": "Endowment",
    "Colaboradoras": "Collaborators",
    "Dirección Misional": "Missional Direction",
    "Administradora de redes sociales": "Social Media Admin",
    "Renta": "Rent",
    "Servicios públicos": "Public Services",
    "Internet": "Internet",
    "Celular": "Mobile",
    "Compras insumos": "Supplies Purchases",
    "Compras equipos de tecnología": "Tech Equipment Purchases",
    "Compras licencias": "Licenses Purchases",
    "Compras inmuebles": "Real Estate Purchases",
    "Obligaciones legales": "Legal Obligations",
    "Insumos cafetería": "Cafeteria Supplies",
    "Insumos de aseo": "Cleaning Supplies",
    "Insumos papelería": "Stationery Supplies",
    "Transporte urbano": "Urban Transport",
    "Transporte colaboradores": "Collaborators Transport",
    "Transporte Parqueaderos": "Parking Transport",
    "Transporte gasolina": "Gasoline Transport",
    "Reparaciones": "Repairs",
    "Mantenimientos": "Maintenance",
    "Reunión Junta": "Board Meeting",
    "Refrigerios colaboradores": "Collaborators Snacks",
    "Facebook": "Facebook",
    "Instagram": "Instagram",
    "Página Web": "Website",
    "Volantes": "Flyers",
    "Botones": "Buttons",
    "Brochure": "Brochure",
    "Pendones / Letrero": "Banners / Signs",
    "Tarjetas de presentación": "Business Cards",
    "Cartas / Sobres con membrete": "Letters / Envelopes",
    "Audiovisuales": "Audiovisuals",
    "Reuniones externas": "External Meetings",
    "Eventos generales": "General Events",
    "Entrega publicidad": "Advertising Delivery",
    "Eventos provida": "Pro-life Events",
    "Gastos bancarios": "Bank Expenses"
}


# ============================
#   RUTAS PRINCIPALES CON ROLES
# ============================

@main_bp.route('/dashboard')
def dashboard():
    if "usuario" not in session:
        return redirect(url_for('index'))
    return render_template('dashboard.html')


@main_bp.route('/madres')
def madres():
    if "usuario" not in session:
        return redirect(url_for('index'))
    return render_template('madres.html')


@main_bp.route('/movimientos')
def movimientos():
    if "usuario" not in session:
        return redirect(url_for('index'))
    return render_template('movimientos.html')


@main_bp.route('/reportes')
def reportes():
    if "usuario" not in session:
        return redirect(url_for('index'))
    return render_template('reportes.html')



@main_bp.route('/api/madres/<int:row_id>', methods=["DELETE"])
def api_delete_madre(row_id):
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    wb = get_workbook()
    sheet = wb["Registro Madres"]

    if row_id < 2 or row_id > sheet.max_row:
        return jsonify({"error": "Madre no encontrada"}), 404

    sheet.delete_rows(row_id, 1)
    save_workbook(wb)
    return jsonify({"message": "Madre eliminada"})

# -------------------------

#  API MOVIMIENTOS (con roles)
# -------------------------
@main_bp.route('/api/movimientos', methods=['GET'])
def api_get_movimientos():
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    tipo_q = (request.args.get('tipo', "") or "").lower()
    medio_q = (request.args.get('medio', "") or "").lower()
    categoria_q = request.args.get('categoria', "") or ""
    subcategoria_q = request.args.get('subcategoria', "") or ""
    codigoMadre_q = (request.args.get('codigoMadre', "") or "").strip()
    fechaDesde_q = request.args.get('fechaDesde', "") or ""
    fechaHasta_q = request.args.get('fechaHasta', "") or ""

    fd = fh = None
    try:
        if fechaDesde_q:
            fd = datetime.fromisoformat(fechaDesde_q).date()
        if fechaHasta_q:
            fh = datetime.fromisoformat(fechaHasta_q).date()
    except Exception:
        pass

    wb = get_workbook(data_only=True)
    results = []

    def check_filters(mov):
        if tipo_q:
            tnorm = mov['tipo'].lower()
            tipo_q_norm = tipo_q[:-1] if tipo_q.endswith('s') else tipo_q
            if tnorm != tipo_q_norm:
                return False
        if medio_q and mov['medio'].lower() != medio_q:
            return False
        if categoria_q and str(mov['categoria']).strip().lower() != str(categoria_q).strip().lower():
            return False
        if subcategoria_q and str(mov['subcategoria']).strip().lower() != str(subcategoria_q).strip().lower():
            return False
        if codigoMadre_q and str(mov['codigoMadre']).strip() != codigoMadre_q:
            return False
        if fd or fh:
            fecha_obj = parse_date_cell(mov['fecha'])
            if not fecha_obj:
                return False
            if fd and fecha_obj < fd:
                return False
            if fh and fecha_obj > fh:
                return False
        return True

    if tipo_q in ("", "ingresos", "ingreso"):
        if "Ingresos" in wb.sheetnames:
            sheet = wb["Ingresos"]
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                mov = row_to_dict(row, i, "Ingresos")
                if check_filters(mov):
                    results.append(mov)

    if tipo_q in ("", "egresos", "egreso"):
        if "Egresos" in wb.sheetnames:
            sheet = wb["Egresos"]
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                mov = row_to_dict(row, i, "Egresos")
                if check_filters(mov):
                    results.append(mov)

    results.sort(key=lambda x: parse_date_cell(x['fecha']) or date.min, reverse=True)
    return jsonify(results)


@main_bp.route('/api/movimientos', methods=['POST'])
def api_post_movimiento():
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    data = request.json or {}
    tipo = str(data.get('tipo', '')).lower()
    tipo_norm = tipo[:-1] if tipo.endswith('s') else tipo

    if tipo_norm not in ('ingreso', 'egreso'):
        return jsonify({"error": "tipo debe ser 'ingreso' o 'egreso'"}), 400

    sheet_name = "Ingresos" if tipo_norm == "ingreso" else "Egresos"

    wb = get_workbook()
    sheet = wb[sheet_name]

    nueva_fila = [
        data.get("fecha"),
        data.get("medio"),
        data.get("tipo"),
        data.get("categoria"),
        data.get("subcategoria"),
        data.get("codigoMadre"),
        data.get("concepto"),
        data.get("valor"),
        data.get("responsable")
    ]
    sheet.append(nueva_fila)

    actualizar_reportes(wb)
    save_workbook(wb)

    return jsonify({"message": "Movimiento agregado", "sheet": sheet_name, "id": sheet.max_row}), 201


@main_bp.route('/api/movimientos/<sheet>/<int:row_id>', methods=['PUT'])
def api_put_movimiento(sheet, row_id):
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    if sheet not in ("Ingresos", "Egresos"):
        return jsonify({"error": "sheet inválido"}), 400

    data = request.json or {}
    wb = get_workbook()
    ws = wb[sheet]

    if row_id < 2 or row_id > ws.max_row:
        return jsonify({"error": "fila no encontrada"}), 404

    try:
        ws.cell(row=row_id, column=1).value = data.get("fecha")
        ws.cell(row=row_id, column=2).value = data.get("medio")
        ws.cell(row=row_id, column=3).value = data.get("tipo")
        ws.cell(row=row_id, column=4).value = data.get("categoria")
        ws.cell(row=row_id, column=5).value = data.get("subcategoria")
        ws.cell(row=row_id, column=6).value = data.get("codigoMadre")
        ws.cell(row=row_id, column=7).value = data.get("concepto")
        ws.cell(row=row_id, column=8).value = data.get("valor")
        ws.cell(row=row_id, column=9).value = data.get("responsable")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    actualizar_reportes(wb)
    save_workbook(wb)
    return jsonify({"message": "Movimiento actualizado"})


@main_bp.route('/api/movimientos/<sheet>/<int:row_id>', methods=['DELETE'])
def api_delete_movimiento(sheet, row_id):
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    if sheet not in ("Ingresos", "Egresos"):
        return jsonify({"error": "sheet inválido"}), 400

    wb = get_workbook()
    ws = wb[sheet]

    if row_id < 2 or row_id > ws.max_row:
        return jsonify({"error": "fila no encontrada"}), 404

    ws.delete_rows(row_id, 1)
    actualizar_reportes(wb)
    save_workbook(wb)
    return jsonify({"message": "Movimiento eliminado"})


# -------------------------
#   API REST - MADRES (con roles)
# -------------------------
@main_bp.route('/api/madres', methods=["GET"])
def get_madres():
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    wb = get_workbook()
    sheet = wb["Registro Madres"]

    data = []
    for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if any(row):
            data.append({
                "id": i,
                "codigoMadre": row[0],
                "nombreCompleto": row[1],
                "identificacion": row[2],
                "fechaIngreso": str(row[3]),
                "procesa": row[4]
            })
    return jsonify(data)


@main_bp.route('/api/madres', methods=["POST"])
def add_madre():
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    data = request.json or {}
    wb = get_workbook()
    sheet = wb["Registro Madres"]

    responsable = session.get("rol")

    nueva_fila = [
        data.get("codigoMadre"),
        data.get("nombreCompleto"),
        data.get("identificacion"),
        data.get("fechaIngreso"),
        responsable
    ]
    sheet.append(nueva_fila)
    save_workbook(wb)

    return jsonify({"message": "Madre registrada con éxito"}), 201


@main_bp.route('/api/madres/<int:row_id>', methods=["PUT"])
def update_madre(row_id):
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401

    if session.get("rol") not in ("admin", "secretaria"):
        return jsonify({"error": "No autorizado"}), 403

    data = request.json or {}
    wb = get_workbook()
    sheet = wb["Registro Madres"]

    if row_id < 2 or row_id > sheet.max_row:
        return jsonify({"error": "Madre no encontrada"}), 404

    try:
        sheet.cell(row=row_id, column=1).value = data.get("codigoMadre")
        sheet.cell(row=row_id, column=2).value = data.get("nombreCompleto")
        sheet.cell(row=row_id, column=3).value = data.get("identificacion")
        sheet.cell(row=row_id, column=4).value = data.get("fechaIngreso")
        sheet.cell(row=row_id, column=5).value = session.get("rol")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    save_workbook(wb)
    return jsonify({"message": "Madre actualizada con éxito"})


# -------------------------
#  SALDOS / RESUMEN
# -------------------------

# ============================
#   API CONFIGURACION (ETAPA 5)
# ============================
def get_config_data(wb):
    if "Configuracion" not in wb.sheetnames:
        return {"saldoInicialCaja": 0, "saldoInicialBanco": 0, "fechaInicio": None}
    
    sheet = wb["Configuracion"]
    config = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0]:
            config[row[0]] = row[1]
    
    fecha_inicio = None
    if config.get("FechaInicioSistema"):
        try:
            from datetime import datetime
            fecha_inicio = datetime.fromisoformat(str(config["FechaInicioSistema"])).date()
        except Exception:
            pass
            
    return {
        "saldoInicialCaja": float(config.get("SaldoInicialCaja", 0)),
        "saldoInicialBanco": float(config.get("SaldoInicialBanco", 0)),
        "fechaInicio": fecha_inicio
    }

@main_bp.route('/api/configuracion', methods=['GET'])
def api_get_config():
    wb = get_workbook(data_only=True)
    config = get_config_data(wb)
    if config["fechaInicio"]:
        config["fechaInicio"] = config["fechaInicio"].isoformat()
    return jsonify(config)

@main_bp.route('/api/configuracion', methods=['POST'])
def api_post_config():
    if "usuario" not in session:
        return jsonify({"error": "No autenticado"}), 401
    if session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403

    data = request.get_json()
    if not data:
        return jsonify({"error": "Datos inválidos"}), 400

    saldo_caja = data.get("saldoInicialCaja", 0)
    saldo_banco = data.get("saldoInicialBanco", 0)
    fecha_inicio_str = data.get("fechaInicio")
    force = data.get("force", False)
    
    try:
        from datetime import datetime
        fecha_inicio = datetime.fromisoformat(fecha_inicio_str).date() if fecha_inicio_str else None
    except Exception:
        return jsonify({"error": "Fecha inválida"}), 400

    wb = get_workbook()
    
    # Validar movimientos anteriores
    if fecha_inicio and not force:
        count_old = 0
        for sheet_name in ["Ingresos", "Egresos"]:
            if sheet_name in wb.sheetnames:
                for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
                    if not any(row): continue
                    f = parse_date_cell(row[0])
                    if f and f < fecha_inicio:
                        count_old += 1
        if count_old > 0:
            return jsonify({
                "warning": True,
                "message": f"Existen {count_old} movimientos anteriores a la fecha seleccionada. Estos no serán tomados en cuenta para los cálculos dinámicos."
            })
    
    if "Configuracion" not in wb.sheetnames:
        wb.create_sheet("Configuracion")
        sheet = wb["Configuracion"]
        sheet.append(["Parametro", "Valor"])
    else:
        sheet = wb["Configuracion"]
        # Limpiar
        if sheet.max_row >= 2:
            sheet.delete_rows(2, sheet.max_row)
            
    sheet.append(["SaldoInicialCaja", saldo_caja])
    sheet.append(["SaldoInicialBanco", saldo_banco])
    if fecha_inicio_str:
        sheet.append(["FechaInicioSistema", fecha_inicio_str])
        
    save_workbook(wb)
    return jsonify({"success": True})


@main_bp.route('/api/saldos', methods=['GET'])
def api_get_saldos():
    wb = get_workbook(data_only=True)
    ingresos_sheet = wb["Ingresos"] if "Ingresos" in wb.sheetnames else None
    egresos_sheet = wb["Egresos"] if "Egresos" in wb.sheetnames else None
    config_data = get_config_data(wb)
    fecha_inicio_sistema = config_data["fechaInicio"]
    
    hoy = datetime.now().date()
    ym = hoy.strftime("%Y-%m")

    saldo_inicial = config_data["saldoInicialBanco"] + config_data["saldoInicialCaja"]
    ingresos_banco_mes = 0.0
    egresos_banco_mes = 0.0
    ingresos_caja_mes = 0.0
    egresos_caja_mes = 0.0
    
    saldo_banco_actual = config_data["saldoInicialBanco"]
    saldo_caja_actual = config_data["saldoInicialCaja"]
    
    total_ingresos_historico = 0.0
    total_egresos_historico = 0.0
    total_movimientos = 0
    
    if ingresos_sheet:
        for row in ingresos_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            fecha = parse_date_cell(row[0])
            if fecha_inicio_sistema and fecha and fecha < fecha_inicio_sistema:
                continue
                
            total_movimientos += 1
            medio = str(row[1]).strip().lower() if row[1] else ""
            valor = 0.0
            try: valor = float(row[7] or 0)
            except: pass
            
            total_ingresos_historico += valor
            
            if medio == "banco":
                saldo_banco_actual += valor
            elif medio == "caja":
                saldo_caja_actual += valor
                
            if fecha:
                f_ym = fecha.strftime("%Y-%m")
                if f_ym < ym:
                    saldo_inicial += valor
                elif f_ym == ym:
                    if medio == "banco": ingresos_banco_mes += valor
                    elif medio == "caja": ingresos_caja_mes += valor

    if egresos_sheet:
        for row in egresos_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            fecha = parse_date_cell(row[0])
            if fecha_inicio_sistema and fecha and fecha < fecha_inicio_sistema:
                continue
                
            total_movimientos += 1
            medio = str(row[1]).strip().lower() if row[1] else ""
            valor = 0.0
            try: valor = float(row[7] or 0)
            except: pass
            
            total_egresos_historico += valor
            
            if medio == "banco":
                saldo_banco_actual -= valor
            elif medio == "caja":
                saldo_caja_actual -= valor
                
            if fecha:
                f_ym = fecha.strftime("%Y-%m")
                if f_ym < ym:
                    saldo_inicial -= valor
                elif f_ym == ym:
                    if medio == "banco": egresos_banco_mes += valor
                    elif medio == "caja": egresos_caja_mes += valor
                    
    return jsonify({
        "saldoInicial": float(saldo_inicial),
        "ingresosBanco": float(ingresos_banco_mes),
        "egresosBanco": float(egresos_banco_mes),
        "saldoBanco": float(saldo_banco_actual),
        "ingresosCaja": float(ingresos_caja_mes),
        "egresosCaja": float(egresos_caja_mes),
        "saldoCaja": float(saldo_caja_actual),
        "saldoConsolidado": float(saldo_banco_actual + saldo_caja_actual),
        
        # Legacy fields for movimientos.js
        "totalIngresos": float(total_ingresos_historico),
        "totalEgresos": float(total_egresos_historico),
        "balance": float(total_ingresos_historico - total_egresos_historico),
        "totalMovimientos": total_movimientos
    })
# -------------------------
#  EXPORTAR MOVIMIENTOS A PDF
# -------------------------
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter, landscape

@main_bp.route('/api/export/pdf', methods=['GET'])
def api_export_pdf_final():
    import io
    from datetime import datetime

    tipo_q = (request.args.get('tipo', "") or "").lower()
    medio_q = (request.args.get('medio', "") or "").lower()
    categoria_q = request.args.get('categoria', "") or ""
    subcategoria_q = request.args.get('subcategoria', "") or ""
    codigoMadre_q = (request.args.get('codigoMadre', "") or "").strip()
    fechaDesde_q = request.args.get('fechaDesde', "") or ""
    fechaHasta_q = request.args.get('fechaHasta', "") or ""

    fd = fh = None
    try:
        if fechaDesde_q:
            fd = datetime.fromisoformat(fechaDesde_q).date()
        if fechaHasta_q:
            fh = datetime.fromisoformat(fechaHasta_q).date()
    except Exception:
        pass

    wb = get_workbook(data_only=True)
    results = []

    def check_filters(mov):
        if tipo_q:
            tnorm = mov['tipo'].lower()
            tipo_q_norm = tipo_q[:-1] if tipo_q.endswith('s') else tipo_q
            if tnorm != tipo_q_norm:
                return False
        if medio_q and mov['medio'].lower() != medio_q:
            return False
        if categoria_q and str(mov['categoria']).strip().lower() != str(categoria_q).strip().lower():
            return False
        if subcategoria_q and str(mov['subcategoria']).strip().lower() != str(subcategoria_q).strip().lower():
            return False
        if codigoMadre_q and str(mov['codigoMadre']).strip() != codigoMadre_q:
            return False
        if fd or fh:
            fecha_obj = parse_date_cell(mov['fecha'])
            if not fecha_obj:
                return False
            if fd and fecha_obj < fd:
                return False
            if fh and fecha_obj > fh:
                return False
        return True

    for sheet_name in ("Ingresos", "Egresos"):
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            for i, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                if not any(row):
                    continue
                mov = row_to_dict(row, i, sheet_name)
                if check_filters(mov):
                    results.append(mov)

    results.sort(key=lambda x: parse_date_cell(x['fecha']) or date.min, reverse=True)

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    elements = []

    styles = getSampleStyleSheet()
    title = Paragraph("Reporte de Movimientos", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 12))

    data = [["Fecha", "Tipo", "Medio", "Categoría", "Subcategoría", "Código Madre", "Concepto", "Valor", "Responsable"]]

    for m in results:
        try:
            valor_fmt = f"${float(m.get('valor', 0)):,.0f}".replace(",", ".")
        except Exception:
            valor_fmt = str(m.get('valor', ''))
        data.append([
            str(m.get("fecha", "")),
            str(m.get("tipo", "")),
            str(m.get("medio", "")),
            str(m.get("categoria", "")),
            str(m.get("subcategoria", "")),
            str(m.get("codigoMadre", "")),
            str(m.get("concepto", "")),
            valor_fmt,
            str(m.get("responsable", "")),
        ])

    table = Table(data, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4CAF50")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 10),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 8),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
    ]))
    elements.append(table)
    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"movimientos_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = make_response(pdf)
    response.headers.set('Content-Type', 'application/pdf')
    response.headers.set('Content-Disposition', f'attachment; filename={filename}')
    return response


# -------------------------
#  API REPORTES
# -------------------------

def get_trm_dict(wb):
    if "TRM" not in wb.sheetnames:
        return {}
    sheet = wb["TRM"]
    res = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if row[0] and row[1]:
            res[str(row[0])] = float(row[1])
    return res

def get_trm_for_date(f_ym, trm_dict):
    if not trm_dict:
        return None, False
    if f_ym in trm_dict:
        return trm_dict[f_ym], False
    
    available = sorted(trm_dict.keys())
    past = [m for m in available if m < f_ym]
    if past:
        return trm_dict[past[-1]], True
    return None, False

# ============================
#   API TRM (ETAPA 6)
# ============================
@main_bp.route('/api/trm', methods=['GET'])
def api_get_trm():
    wb = get_workbook(data_only=True)
    return jsonify(get_trm_dict(wb))

@main_bp.route('/api/trm', methods=['POST'])
def api_post_trm():
    if "usuario" not in session or session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403
    data = request.get_json()
    mes = data.get("mes")
    valor = data.get("valor")
    if not mes or not valor:
        return jsonify({"error": "Datos inválidos"}), 400
        
    wb = get_workbook()
    if "TRM" not in wb.sheetnames:
        wb.create_sheet("TRM")
        sheet = wb["TRM"]
        sheet.append(["Mes", "Valor"])
    else:
        sheet = wb["TRM"]
        
    found = False
    for row in sheet.iter_rows(min_row=2):
        if row[0].value == mes:
            row[1].value = float(valor)
            found = True
            break
            
    if not found:
        sheet.append([mes, float(valor)])
        
    save_workbook(wb)
    return jsonify({"success": True})

@main_bp.route('/api/trm/<mes>', methods=['DELETE'])
def api_delete_trm(mes):
    if "usuario" not in session or session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403
        
    wb = get_workbook()
    if "TRM" not in wb.sheetnames:
        return jsonify({"error": "No hay TRM"}), 404
        
    sheet = wb["TRM"]
    for i, row in enumerate(sheet.iter_rows(min_row=2), start=2):
        if str(row[0].value) == mes:
            sheet.delete_rows(i, 1)
            save_workbook(wb)
            return jsonify({"success": True})
            
    return jsonify({"error": "Mes no encontrado"}), 404

@main_bp.route('/api/reportes', methods=['GET'])
def api_get_reportes():
    wb = get_workbook(data_only=True)
    ingresos_sheet = wb["Ingresos"] if "Ingresos" in wb.sheetnames else None
    egresos_sheet = wb["Egresos"] if "Egresos" in wb.sheetnames else None
    config_data = get_config_data(wb)
    fecha_inicio_sistema = config_data['fechaInicio']


    tipo_reporte = request.args.get("tipoReporte", "mensual")
    anio = request.args.get("anio", "")
    mes = request.args.get("mes", "all")

    
    moneda = request.args.get("moneda", "cop").lower()
    trm_dict = get_trm_dict(wb)
    trm_warnings = set()
    
    def convert_usd(fecha, val):
        if moneda == "cop": return val
        if isinstance(fecha, str):
            f_ym = fecha[:7]
        else:
            f_ym = fecha.strftime("%Y-%m")
        trm, is_fb = get_trm_for_date(f_ym, trm_dict)
        if not trm:
            raise ValueError(f"Falta configurar la TRM para el mes {f_ym} o anterior.")
        if is_fb:
            trm_warnings.add(f"{f_ym} (usó {trm})")
        return val / trm
        
    try:
        config_data["saldoInicialBanco"] = convert_usd(fecha_inicio_sistema, config_data["saldoInicialBanco"]) if fecha_inicio_sistema else config_data["saldoInicialBanco"]
        config_data["saldoInicialCaja"] = convert_usd(fecha_inicio_sistema, config_data["saldoInicialCaja"]) if fecha_inicio_sistema else config_data["saldoInicialCaja"]
    except ValueError as e:
        return jsonify({"error": str(e), "need_trm": True}), 400
    ingresos_totales = 0.0
    egresos_totales = 0.0

    evolucion_mensual = {}
    distribucion_gastos = {}
    comparativo_anual = {}
    
    # Recopilar todos los movimientos cronológicamente para la nueva tabla
    movs = []

    def get_month_label(date_obj):
        return date_obj.strftime("%Y-%m")
        
    def get_year_label(date_obj):
        return str(date_obj.year)

    if ingresos_sheet:
        for row in ingresos_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            f = parse_date_cell(row[0])
            if not f: continue
            if fecha_inicio_sistema and f < fecha_inicio_sistema: continue
            m = str(row[1]).strip().lower() if row[1] else ""
            v = 0.0
            try: v = float(row[7] or 0)
            except: pass
            try: v = convert_usd(f, v)
            except ValueError as e: return jsonify({"error": str(e), "need_trm": True}), 400
            
            movs.append({"fecha": f, "tipo": "ingreso", "medio": m, "valor": v})
            
            # Filtros para tarjetas y gráficas
            match = True
            if anio and str(f.year) != anio: match = False
            if mes != "all" and str(f.month).zfill(2) != mes: match = False
            
            if match:
                ingresos_totales += v
                mlbl = get_month_label(f)
                ylbl = get_year_label(f)
                
                if mlbl not in evolucion_mensual:
                    evolucion_mensual[mlbl] = {"ingresos": 0, "egresos": 0}
                evolucion_mensual[mlbl]["ingresos"] += v
                
                if ylbl not in comparativo_anual:
                    comparativo_anual[ylbl] = {"ingresos": 0, "egresos": 0}
                comparativo_anual[ylbl]["ingresos"] += v

    if egresos_sheet:
        for row in egresos_sheet.iter_rows(min_row=2, values_only=True):
            if not any(row): continue
            f = parse_date_cell(row[0])
            if not f: continue
            if fecha_inicio_sistema and f < fecha_inicio_sistema: continue
            m = str(row[1]).strip().lower() if row[1] else ""
            cat = str(row[3]).strip() if row[3] else "Otros"
            v = 0.0
            try: v = float(row[7] or 0)
            except: pass
            try: v = convert_usd(f, v)
            except ValueError as e: return jsonify({"error": str(e), "need_trm": True}), 400
            
            movs.append({"fecha": f, "tipo": "egreso", "medio": m, "valor": v, "cat": cat})
            
            # Filtros para tarjetas y gráficas
            match = True
            if anio and str(f.year) != anio: match = False
            if mes != "all" and str(f.month).zfill(2) != mes: match = False
            
            if match:
                egresos_totales += v
                mlbl = get_month_label(f)
                ylbl = get_year_label(f)
                
                if mlbl not in evolucion_mensual:
                    evolucion_mensual[mlbl] = {"ingresos": 0, "egresos": 0}
                evolucion_mensual[mlbl]["egresos"] += v
                
                if ylbl not in comparativo_anual:
                    comparativo_anual[ylbl] = {"ingresos": 0, "egresos": 0}
                comparativo_anual[ylbl]["egresos"] += v
                
                distribucion_gastos[cat] = distribucion_gastos.get(cat, 0) + v

    # Calcular madres atendidas (total histórico o según filtro)
    madres_atendidas = 0
    madres_file = Config.EXCEL_FILE
    if os.path.exists(madres_file):
        try:
            wb_m = openpyxl.load_workbook(madres_file, data_only=True)
            if "Madres" in wb_m.sheetnames:
                ws_m = wb_m["Madres"]
                for row in ws_m.iter_rows(min_row=2, values_only=True):
                    if row[0]: madres_atendidas += 1
        except: pass

    # --- Lógica de Tabla Cronológica (Nueva Etapa 4) ---
    movs.sort(key=lambda x: x["fecha"])
    from collections import OrderedDict
    meses_data = OrderedDict()
    
    running_banco = config_data['saldoInicialBanco']
    running_caja = config_data['saldoInicialCaja']
    running_consolidado = config_data['saldoInicialBanco'] + config_data['saldoInicialCaja']
    
    for mov in movs:
        ym = mov["fecha"].strftime("%Y-%m")
        if ym not in meses_data:
            meses_data[ym] = {
                "saldoInicial": running_consolidado,
                "ing_banco": 0.0, "egr_banco": 0.0,
                "ing_caja": 0.0, "egr_caja": 0.0
            }
            
        v = mov["valor"]
        if mov["tipo"] == "ingreso":
            if mov["medio"] == "banco":
                meses_data[ym]["ing_banco"] += v
                running_banco += v
            elif mov["medio"] == "caja":
                meses_data[ym]["ing_caja"] += v
                running_caja += v
            running_consolidado += v
        else:
            if mov["medio"] == "banco":
                meses_data[ym]["egr_banco"] += v
                running_banco -= v
            elif mov["medio"] == "caja":
                meses_data[ym]["egr_caja"] += v
                running_caja -= v
            running_consolidado -= v
            
        meses_data[ym]["saldoBanco"] = running_banco
        meses_data[ym]["saldoCaja"] = running_caja
        meses_data[ym]["saldoFinal"] = running_consolidado

    tabla_cronologica = []
    for ym, d in meses_data.items():
        tabla_cronologica.append({
            "mes": ym,
            "saldoInicial": d["saldoInicial"],
            "ingresosBancos": d["ing_banco"],
            "egresosBancos": d["egr_banco"],
            "saldoBancos": d["saldoBanco"],
            "ingresosCaja": d["ing_caja"],
            "egresosCaja": d["egr_caja"],
            "saldoCaja": d["saldoCaja"],
            "saldoFinal": d["saldoFinal"]
        })

    lang = request.args.get("lang", "es").lower()
    if lang == "en":
        distribucion_gastos = {CAT_EN.get(k, k): v for k, v in distribucion_gastos.items()}

    return jsonify({
        'trmWarnings': list(trm_warnings),
        "ingresosTotales": ingresos_totales,
        "egresosTotales": egresos_totales,
        "balanceNeto": ingresos_totales - egresos_totales,
        "madresAtendidas": madres_atendidas,
        "evolucionMensual": evolucion_mensual,
        "distribucionGastos": distribucion_gastos,
        "comparativoAnual": comparativo_anual,
        "tablaMensual": tabla_cronologica
    })
# ===========================================================
#  ✅ EXPORT REPORTES (PDF y EXCEL) CON MESES + CONSOLIDADO
#  ✅ + RESUMEN POR MADRE (con NOMBRE)
#  ✅ + IDIOMA (ES/EN) via ?lang=es|en
# ===========================================================
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
    5: "Mayo", 6: "Junio", 7: "Julio", 8: "Agosto",
    9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre"
}
MESES_EN = {
    1: "January", 2: "February", 3: "March", 4: "April",
    5: "May", 6: "June", 7: "July", 8: "August",
    9: "September", 10: "October", 11: "November", 12: "December"
}
MESES_ORDEN = list(range(1, 13))


# ==========================
#  🌍 TRADUCCIÓN TIPO / MEDIO
# ==========================
TIPO_EN = {
    "ingreso": "income",
    "egreso": "expense",
}

MEDIO_EN = {
    "banco": "bank",
    "caja": "cash",
}

def traducir_tipo(valor, lang):
    if lang != "en":
        return valor
    v = (str(valor) or "").strip().lower()
    return TIPO_EN.get(v, valor)

def traducir_medio(valor, lang):
    if lang != "en":
        return valor
    v = (str(valor) or "").strip().lower()
    return MEDIO_EN.get(v, valor)



def _tr_label(value, lang, dic_en):
    """
    Traduce ES → EN si lang == 'en'
    Si no existe traducción, deja el texto original
    """
    if value is None:
        return ""
    s = str(value).strip()
    if lang == "en":
        return dic_en.get(s, s)
    return s



# ==========================
#  🌍 TRADUCCIONES ES / EN
# ==========================
TEXTOS = {
    "es": {
        "informe_titulo": "INFORME MENSUAL (Enero - Diciembre) + CONSOLIDADO",
        "fecha_corte": "Fecha de corte",
        "ingresos": "INGRESOS",
        "egresos": "EGRESOS",
        "resumen": "RESUMEN",
        "resumen_madre": "RESUMEN POR MADRE",
        "detalle_madre": "DETALLE POR MADRE",
        "codigo_madre": "Código Madre",
        "nombre": "Nombre",
        "balance": "Balance",
        "consolidado_ing": "Consolidado Ingresos",
        "consolidado_egr": "Consolidado Egresos",
        "sin_datos": "(Sin datos)",
        "sin_movimientos": "(Sin movimientos con código madre)",
        "categoria": "Categoría",
        "subcategoria": "Subcategoría",
        "fecha": "Fecha",
        "tipo": "Tipo",
        "medio": "Medio",
        "concepto": "Concepto",
        "valor": "Valor",
        "categoria_sub": "Categoría / Subcategoría",
        "consolidado": "Consolidado"
    },
    "en": {
        "informe_titulo": "MONTHLY REPORT (January - December) + CONSOLIDATED",
        "fecha_corte": "Cut-off date",
        "ingresos": "INCOME",
        "egresos": "EXPENSES",
        "resumen": "SUMMARY",
        "resumen_madre": "SUMMARY BY MOTHER",
        "detalle_madre": "DETAIL BY MOTHER",
        "codigo_madre": "Mother Code",
        "nombre": "Name",
        "balance": "Balance",
        "consolidado_ing": "Total Income",
        "consolidado_egr": "Total Expenses",
        "sin_datos": "(No data)",
        "sin_movimientos": "(No movements with mother code)",
        "categoria": "Category",
        "subcategoria": "Subcategory",
        "fecha": "Date",
        "tipo": "Type",
        "medio": "Method",
        "concepto": "Description",
        "valor": "Amount",
        "categoria_sub": "Category / Subcategory",
        "consolidado": "Consolidated"
    }
}


def _to_date_safe(s):
    if not s:
        return None
    try:
        return datetime.fromisoformat(s).date()
    except:
        try:
            return parse_date_cell(s)
        except:
            return None


def _normalizar_df(df, tipo_fijo):
    expected_cols = ["fecha", "medio", "tipo", "categoria", "subcategoria",
                     "codigoMadre", "concepto", "valor", "responsable"]
    df = df[df.columns[:len(expected_cols)]].copy()
    df.columns = expected_cols
    df["tipo"] = tipo_fijo
    return df


def _aplicar_filtros_reportes(df):
    tipo = (request.args.get("tipo", "") or "").lower().rstrip("s")
    medio = (request.args.get("medio", "") or "").strip().lower()
    madre = (request.args.get("madre", "") or "").strip()

    categorias = request.args.get("categorias", "")
    subcategorias = request.args.get("subcategorias", "")
    categorias = [c.strip() for c in categorias.split(",") if c.strip()]
    subcategorias = [s.strip() for s in subcategorias.split(",") if s.strip()]

    fd = _to_date_safe(request.args.get("fechaDesde", "") or "")
    fh = _to_date_safe(request.args.get("fechaHasta", "") or "")

    if tipo:
        df = df[df["tipo"].astype(str).str.lower() == tipo]
    if medio:
        df = df[df["medio"].astype(str).str.lower() == medio]
    if madre:
        df = df[df["codigoMadre"].astype(str).str.strip() == madre]
    if categorias:
        df = df[df["categoria"].isin(categorias)]
    if subcategorias:
        df = df[df["subcategoria"].isin(subcategorias)]

    if fd:
        df = df[df["fecha"].apply(parse_date_cell) >= fd]
    if fh:
        df = df[df["fecha"].apply(parse_date_cell) <= fh]

    return df, fd, fh


def _construir_tabla_mensual(df, lang="es"):
    estructura = {}
    if df.empty:
        return estructura

    df = df.copy()
    df["fecha_dt"] = df["fecha"].apply(parse_date_cell)
    df = df[df["fecha_dt"].notna()]
    if df.empty:
        return estructura

    df["mes"] = df["fecha_dt"].apply(lambda d: int(d.month))
    df["categoria"] = df["categoria"].fillna("Sin categoría").astype(str)
    df["subcategoria"] = df["subcategoria"].fillna("Sin subcategoría").astype(str)
    df["valor"] = df["valor"].fillna(0).astype(float)

    # ✅ Traducción automática si es inglés
    if lang == "en":
        df["categoria"] = df["categoria"].apply(lambda x: _tr_label(x, lang, CAT_EN))
        df["subcategoria"] = df["subcategoria"].apply(lambda x: _tr_label(x, lang, CAT_EN))

    grp = (
        df.groupby(["categoria", "subcategoria", "mes"], dropna=False)["valor"]
        .sum()
        .reset_index()
    )

    for _, row in grp.iterrows():
        cat = row["categoria"]
        sub = row["subcategoria"]
        mes = int(row["mes"])
        val = float(row["valor"])

        if cat not in estructura:
            estructura[cat] = {}
        if sub not in estructura[cat]:
            estructura[cat][sub] = {m: 0.0 for m in MESES_ORDEN}
            estructura[cat][sub]["consolidado"] = 0.0

        estructura[cat][sub][mes] += val

    for cat, subs in estructura.items():
        for sub, meses_vals in subs.items():
            meses_vals["consolidado"] = sum(meses_vals[m] for m in MESES_ORDEN)

    return estructura



def _money_str(v):
    try:
        return f"${float(v):,.0f}".replace(",", ".")
    except:
        return "$0"


def _money_num(v):
    try:
        return float(v)
    except:
        return 0.0


# ✅ FIX: Normalizar códigos madre (202020.0 -> 202020)
def _norm_cod_madre(x):
    if x is None:
        return ""
    s = str(x).strip()
    if not s or s.lower() == "nan":
        return ""
    try:
        f = float(s)
        if f.is_integer():
            return str(int(f))
    except:
        pass
    if s.endswith(".0"):
        s2 = s[:-2]
        if s2.isdigit():
            return s2
    return s


def _leer_mapa_madres():
    mapa = {}
    try:
        dfm = pd.read_excel(Config.EXCEL_FILE, sheet_name="Registro Madres")
        dfm.columns = [str(c).strip().lower() for c in dfm.columns]
        col_codigo = next((c for c in dfm.columns if "codigo" in c and "madre" in c), None)
        col_nombre = next((c for c in dfm.columns if "nombre" in c), None)
        if not col_codigo or not col_nombre:
            return mapa
        for _, r in dfm.iterrows():
            cod = _norm_cod_madre(r.get(col_codigo, ""))
            nom = str(r.get(col_nombre, "")).strip()
            if cod:
                mapa[cod] = "" if nom.lower() == "nan" else nom
    except:
        pass
    return mapa


def _armar_madres_resumen_y_detalle(df_ing, df_egr):
    mapa = _leer_mapa_madres()

    def prep(df):
        if df is None or df.empty:
            return pd.DataFrame(columns=[
                "fecha", "tipo", "medio", "categoria", "subcategoria",
                "codigoMadre", "concepto", "valor", "responsable",
                "nombre", "fecha_dt"
            ])
        d = df.copy()
        d["codigoMadre"] = d["codigoMadre"].apply(_norm_cod_madre)
        d = d[d["codigoMadre"].astype(str).str.strip() != ""].copy()
        d["nombre"] = d["codigoMadre"].apply(lambda c: mapa.get(c, ""))
        d["fecha_dt"] = d["fecha"].apply(parse_date_cell)
        d = d.sort_values(by=["codigoMadre", "fecha_dt"], ascending=[True, True])
        return d

    di = prep(df_ing)
    de = prep(df_egr)

    resumen_rows = []
    codigos = sorted(set(di["codigoMadre"].dropna().tolist()) | set(de["codigoMadre"].dropna().tolist()))
    codigos = [c for c in codigos if c]

    for cod in codigos:
        nom = mapa.get(cod, "")
        ing = float(di.loc[di["codigoMadre"] == cod, "valor"].sum()) if not di.empty else 0.0
        egr = float(de.loc[de["codigoMadre"] == cod, "valor"].sum()) if not de.empty else 0.0
        bal = ing - egr
        if ing == 0 and egr == 0:
            continue
        resumen_rows.append({
            "codigoMadre": cod,
            "nombre": nom,
            "ingresos": ing,
            "egresos": egr,
            "balance": bal
        })

    resumen_df = pd.DataFrame(resumen_rows)

    det = pd.concat([di, de], ignore_index=True)
    if not det.empty:
        det["tipo"] = det["tipo"].astype(str).str.lower()
        det["valor"] = det["valor"].fillna(0).astype(float)
        det["fecha_dt"] = det["fecha"].apply(parse_date_cell)
        det = det.sort_values(by=["codigoMadre", "fecha_dt", "tipo"], ascending=[True, True, True])

    return resumen_df, det


# -------------------------
#  EXPORT REPORTES PDF
# -------------------------
@main_bp.route('/api/reportes/export/pdf', methods=['GET'])
def api_export_reportes_pdf_mensual_consolidado():
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.pagesizes import landscape, letter

    # ✅ idioma
    lang = (request.args.get("lang", "es") or "es").lower()
    if lang not in TEXTOS:
        lang = "es"
    T = TEXTOS[lang]
    MESES = MESES_EN if lang == "en" else MESES_ES
    # ✅ moneda & TRM
    moneda = request.args.get("moneda", "cop").lower()
    trm_dict = get_trm_dict(get_workbook(data_only=True))
    trm_warnings = set()
    def _convert(fecha, val):
        if moneda == "cop": return val
        f_ym = fecha.strftime("%Y-%m") if not isinstance(fecha, str) else fecha[:7]
        trm, is_fb = get_trm_for_date(f_ym, trm_dict)
        if not trm: raise ValueError(f"Falta configurar la TRM para el mes {f_ym} o anterior.")
        if is_fb: trm_warnings.add(f"{f_ym} (usó {trm})")
        return val / trm

    # ✅ moneda & TRM
    moneda = request.args.get("moneda", "cop").lower()
    trm_dict = get_trm_dict(get_workbook(data_only=True))
    trm_warnings = set()
    def _convert(fecha, val):
        if moneda == "cop": return val
        if not fecha: raise ValueError("Fecha inválida en un registro.")
        f_ym = fecha.strftime("%Y-%m") if not isinstance(fecha, str) else fecha[:7]
        trm, is_fb = get_trm_for_date(f_ym, trm_dict)
        if not trm: raise ValueError(f"Falta configurar la TRM para el mes {f_ym} o anterior.")
        if is_fb: trm_warnings.add(f"{f_ym} (usó {trm})")
        return val / trm


    try:
        df_ing = pd.read_excel(Config.EXCEL_FILE, sheet_name="Ingresos")
    except:
        df_ing = pd.DataFrame()

    try:
        df_egr = pd.read_excel(Config.EXCEL_FILE, sheet_name="Egresos")
    except:
        df_egr = pd.DataFrame()


    try:
        if not df_ing.empty:
            df_ing = _normalizar_df(df_ing, "ingreso")
            df_ing["valor"] = df_ing.apply(lambda r: _convert(parse_date_cell(r["fecha"]), r["valor"]), axis=1)
        if not df_egr.empty:
            df_egr = _normalizar_df(df_egr, "egreso")
            df_egr["valor"] = df_egr.apply(lambda r: _convert(parse_date_cell(r["fecha"]), r["valor"]), axis=1)
    except ValueError as e:
        return jsonify({"error": str(e), "need_trm": True}), 400


    df_ing, fd, fh = _aplicar_filtros_reportes(df_ing) if not df_ing.empty else (df_ing, None, None)
    df_egr, fd2, fh2 = _aplicar_filtros_reportes(df_egr) if not df_egr.empty else (df_egr, None, None)
    fd = fd or fd2
    fh = fh or fh2

    estructura_ing = _construir_tabla_mensual(df_ing, lang=lang) if not df_ing.empty else {}
    estructura_egr = _construir_tabla_mensual(df_egr, lang=lang) if not df_egr.empty else {}


    resumen_madres_df, detalle_madres_df = _armar_madres_resumen_y_detalle(df_ing, df_egr)

    ref_year = datetime.now().year
    if fh:
        ref_year = fh.year
    elif fd:
        ref_year = fd.year
    else:
        max_y = None
        for df_ in (df_ing, df_egr):
            if not df_.empty and 'fecha' in df_.columns:
                try:
                    df_temp = pd.to_datetime(df_['fecha'], errors='coerce')
                    m_y = df_temp.max().year
                    if pd.notna(m_y):
                        max_y = max(max_y or 0, int(m_y))
                except:
                    pass
        if max_y:
            ref_year = max_y

    meses_headers = []
    for m in MESES_ORDEN:
        if moneda == "usd":
            ym = f"{ref_year}-{m:02d}"
            try:
                trm_val, _ = get_trm_for_date(ym, trm_dict)
                if trm_val:
                    meses_headers.append(f"{MESES[m]}<br/>(TRM {int(trm_val)})")
                else:
                    meses_headers.append(MESES[m])
            except:
                meses_headers.append(MESES[m])
        else:
            meses_headers.append(MESES[m])

    # Convertir a Paragraphs para que soporten salto de línea y tengan el ancho adecuado
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.enums import TA_CENTER, TA_LEFT
    from reportlab.platypus import Paragraph
    from reportlab.lib import colors
    
    styles = getSampleStyleSheet()
    styleH = styles['Normal'].clone('HeaderStyle')
    styleH.fontName = 'Helvetica-Bold'
    styleH.fontSize = 7
    styleH.textColor = colors.white
    styleH.alignment = TA_CENTER

    header = [Paragraph(T["categoria_sub"], styleH)] + [Paragraph(h, styleH) for h in meses_headers] + [Paragraph(T["consolidado"], styleH)]

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=18, rightMargin=18, topMargin=18, bottomMargin=18
    )
    styles = getSampleStyleSheet()
    elements = []

    # General
    elements.append(Paragraph(f"<b>{T['informe_titulo']}</b>", styles["Title"]))
    corte = fh.strftime("%d/%m/%Y") if fh else datetime.now().strftime("%d/%m/%Y")
    elements.append(Paragraph(f"<b>{T['fecha_corte']}:</b> {corte}", styles["Normal"]))
    elements.append(Spacer(1, 10))

    def tabla_seccion(titulo, estructura, color_hex):
        elements.append(Paragraph(f"<b>{titulo}</b>", styles["Heading2"]))
        elements.append(Spacer(1, 6))

        data = [header]

        if not estructura:
            data.append([T["sin_datos"]] + [""] * 12 + [""])
        else:
            for cat in sorted(estructura.keys()):
                data.append([f"■ {cat}"] + [""] * 12 + [""])

                subs = estructura[cat]
                for sub in sorted(subs.keys()):
                    fila = [f"   • {sub}"]
                    for m in MESES_ORDEN:
                        v = subs[sub][m]
                        fila.append(_money_str(v) if v else "")
                    fila.append(_money_str(subs[sub]["consolidado"]) if subs[sub]["consolidado"] else "")
                    data.append(fila)

                total_cat = {m: 0.0 for m in MESES_ORDEN}
                for sub in subs:
                    for m in MESES_ORDEN:
                        total_cat[m] += subs[sub][m]
                total_con = sum(total_cat[m] for m in MESES_ORDEN)

                fila_total = [f"TOTAL {cat}"]
                for m in MESES_ORDEN:
                    fila_total.append(_money_str(total_cat[m]) if total_cat[m] else "")
                fila_total.append(_money_str(total_con) if total_con else "")
                data.append(fila_total)

        table = Table(data, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor(color_hex)),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 8),
            ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("FONTSIZE", (0, 1), (-1, -1), 7),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("BOTTOMPADDING", (0, 0), (-1, 0), 6),
        ]))

        for i in range(1, len(data)):
            if str(data[i][0]).startswith("TOTAL "):
                table.setStyle(TableStyle([
                    ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                    ("BACKGROUND", (0, i), (-1, i), colors.whitesmoke),
                ]))

        elements.append(table)
        elements.append(Spacer(1, 12))

    tabla_seccion(T["ingresos"], estructura_ing, "#C8102E")
    tabla_seccion(T["egresos"], estructura_egr, "#C8102E")

    total_ing = float(df_ing["valor"].sum()) if not df_ing.empty else 0.0
    total_egr = float(df_egr["valor"].sum()) if not df_egr.empty else 0.0
    balance = total_ing - total_egr

    elements.append(Paragraph(f"<b>{T['resumen']}</b>", styles["Heading2"]))
    resumen = [
        [T["consolidado_ing"], _money_str(total_ing)],
        [T["consolidado_egr"], _money_str(total_egr)],
        [T["balance"], _money_str(balance)]
    ]
    t_res = Table(resumen, colWidths=[260, 140])
    t_res.setStyle(TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("BACKGROUND", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, -1), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "RIGHT"),
    ]))
    elements.append(t_res)
    elements.append(Spacer(1, 14))

    # Resumen por madre
    elements.append(Paragraph(f"<b>{T['resumen_madre']}</b>", styles["Heading2"]))
    if resumen_madres_df is None or resumen_madres_df.empty:
        elements.append(Paragraph(T["sin_movimientos"], styles["Normal"]))
    else:
        data_m = [[T["codigo_madre"], T["nombre"], T["ingresos"], T["egresos"], T["balance"]]]
        for _, r in resumen_madres_df.iterrows():
            data_m.append([
                str(r.get("codigoMadre", "")),
                str(r.get("nombre", "")),
                _money_str(r.get("ingresos", 0)),
                _money_str(r.get("egresos", 0)),
                _money_str(r.get("balance", 0)),
            ])

        t_m = Table(data_m, repeatRows=1, colWidths=[90, 210, 90, 90, 90])
        t_m.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#263238")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
            ("ALIGN", (2, 1), (-1, -1), "RIGHT"),
            ("ALIGN", (0, 0), (1, -1), "LEFT"),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        elements.append(t_m)

    # Detalle por madre
    elements.append(PageBreak())
    elements.append(Paragraph(f"<b>{T['detalle_madre']}</b>", styles["Title"]))
    elements.append(Spacer(1, 10))

    if detalle_madres_df is None or detalle_madres_df.empty:
        elements.append(Paragraph(T["sin_datos"], styles["Normal"]))
    else:
        for cod, subdf in detalle_madres_df.groupby("codigoMadre"):
            try:
                nombre = str(subdf["nombre"].iloc[0] or "").strip()
            except:
                nombre = ""

            titulo_m = f"{cod} - {nombre}" if nombre else f"{cod} -"
            elements.append(Paragraph(f"<b>{titulo_m}</b>", styles["Heading2"]))
            elements.append(Spacer(1, 6))

            data_det = [[T["fecha"], T["tipo"], T["medio"], T["categoria"], T["subcategoria"], T["concepto"], T["valor"]]]
            for _, rr in subdf.iterrows():
                f = rr.get("fecha_dt", None)
                f_txt = f.strftime("%d/%m/%Y") if isinstance(f, (datetime, date)) and f else str(rr.get("fecha", ""))
                data_det.append([
                    f_txt,
                    traducir_tipo(rr.get("tipo", ""), lang),
                    traducir_medio(rr.get("medio", ""), lang),
                    _tr_label(rr.get("categoria", ""), lang, CAT_EN),
                    _tr_label(rr.get("subcategoria", ""), lang, CAT_EN),
                    str(rr.get("concepto", "")),
                    _money_str(rr.get("valor", 0)),
                ])

            t_d = Table(data_det, repeatRows=1, colWidths=[70, 55, 55, 140, 140, 180, 70])
            t_d.setStyle(TableStyle([
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#455A64")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                ("ALIGN", (6, 1), (6, -1), "RIGHT"),
                ("ALIGN", (0, 0), (-2, -1), "LEFT"),
                ("FONTSIZE", (0, 0), (-1, -1), 7),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ]))
            elements.append(t_d)
            elements.append(Spacer(1, 12))

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    filename = f"Informe_Mensual_Consolidado_{lang.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
    response = make_response(pdf)
    response.headers.set("Content-Type", "application/pdf")
    response.headers.set("Content-Disposition", f"attachment; filename={filename}")
    return response


# -------------------------
#  EXPORT REPORTES EXCEL
# -------------------------
@main_bp.route('/api/reportes/export/excel', methods=['GET'])
def api_export_reportes_excel_mensual_consolidado():
    from openpyxl import Workbook

    # ✅ idioma
    lang = (request.args.get("lang", "es") or "es").lower()
    if lang not in TEXTOS:
        lang = "es"
    T = TEXTOS[lang]
    MESES = MESES_EN if lang == "en" else MESES_ES

    try:
        df_ing = pd.read_excel(Config.EXCEL_FILE, sheet_name="Ingresos")
    except:
        df_ing = pd.DataFrame()

    try:
        df_egr = pd.read_excel(Config.EXCEL_FILE, sheet_name="Egresos")
    except:
        df_egr = pd.DataFrame()


    if not df_ing.empty:
        df_ing = _normalizar_df(df_ing, "ingreso")
        df_ing["valor"] = df_ing.apply(lambda r: _convert(parse_date_cell(r["fecha"]), r["valor"]), axis=1)
    if not df_egr.empty:
        df_egr = _normalizar_df(df_egr, "egreso")
        df_egr["valor"] = df_egr.apply(lambda r: _convert(parse_date_cell(r["fecha"]), r["valor"]), axis=1)


    df_ing, fd, fh = _aplicar_filtros_reportes(df_ing) if not df_ing.empty else (df_ing, None, None)
    df_egr, fd2, fh2 = _aplicar_filtros_reportes(df_egr) if not df_egr.empty else (df_egr, None, None)
    fd = fd or fd2
    fh = fh or fh2

    estructura_ing = _construir_tabla_mensual(df_ing, lang=lang) if not df_ing.empty else {}
    estructura_egr = _construir_tabla_mensual(df_egr, lang=lang) if not df_egr.empty else {}


    resumen_madres_df, detalle_madres_df = _armar_madres_resumen_y_detalle(df_ing, df_egr)

    ref_year = datetime.now().year
    if fh:
        ref_year = fh.year
    elif fd:
        ref_year = fd.year
    else:
        max_y = None
        for df_ in (df_ing, df_egr):
            if not df_.empty and 'fecha' in df_.columns:
                try:
                    df_temp = pd.to_datetime(df_['fecha'], errors='coerce')
                    m_y = df_temp.max().year
                    if pd.notna(m_y):
                        max_y = max(max_y or 0, int(m_y))
                except:
                    pass
        if max_y:
            ref_year = max_y

    meses_headers = []
    for m in MESES_ORDEN:
        if moneda == "usd":
            ym = f"{ref_year}-{m:02d}"
            try:
                trm_val, _ = get_trm_for_date(ym, trm_dict)
                if trm_val:
                    meses_headers.append(f"{MESES[m]}\n(TRM {int(trm_val)})")
                else:
                    meses_headers.append(MESES[m])
            except:
                meses_headers.append(MESES[m])
        else:
            meses_headers.append(MESES[m])

    headers = [T["categoria_sub"]] + meses_headers + [T["consolidado"]]

    wb_out = Workbook()

    ws = wb_out.active
    ws.title = "General Report" if lang == "en" else "Informe General"

    title_font = Font(bold=True, size=14)
    bold = Font(bold=True)
    white_bold = Font(bold=True, color="FFFFFF")
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    fill_red = PatternFill(start_color="C8102E", end_color="C8102E", fill_type="solid")
    fill_gray = PatternFill(start_color="EDEDED", end_color="EDEDED", fill_type="solid")
    fill_light = PatternFill(start_color="F7F7F7", end_color="F7F7F7", fill_type="solid")
    fill_dark = PatternFill(start_color="263238", end_color="263238", fill_type="solid")

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws["A1"] = T["informe_titulo"]
    ws["A1"].font = title_font
    ws.merge_cells("A1:N1")

    corte = fh.strftime("%d/%m/%Y") if fh else datetime.now().strftime("%d/%m/%Y")
    ws["A2"] = f"{T['fecha_corte']}:"
    ws["B2"] = corte
    ws["A2"].font = bold
    ws["A2"].alignment = align_left
    ws["B2"].alignment = align_left

    row = 4
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=h)
        cell.font = white_bold
        cell.fill = fill_red
        cell.alignment = align_left if col == 1 else align_right
        cell.border = border
    row += 1

    def escribir_seccion(nombre, estructura):
        nonlocal row

        ws.cell(row=row, column=1, value=nombre).font = white_bold
        for col in range(1, len(headers) + 1):
            c = ws.cell(row=row, column=col)
            c.fill = fill_red
            c.border = border
            c.alignment = align_left if col == 1 else align_right
        row += 1

        if not estructura:
            ws.cell(row=row, column=1, value=T["sin_datos"]).alignment = align_left
            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).border = border
            row += 2
            return

        for cat in sorted(estructura.keys()):
            ws.cell(row=row, column=1, value=cat).font = bold
            for col in range(1, len(headers) + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = fill_gray
                cell.border = border
                cell.alignment = align_left if col == 1 else align_right
            row += 1

            subs = estructura[cat]

            for sub in sorted(subs.keys()):
                ws.cell(row=row, column=1, value=f"  • {sub}").alignment = align_left
                ws.cell(row=row, column=1).border = border

                for idx, m in enumerate(MESES_ORDEN, start=2):
                    v = subs[sub][m]
                    cell = ws.cell(row=row, column=idx, value=_money_num(v) if v else "")
                    cell.number_format = '#,##0'
                    cell.alignment = align_right
                    cell.border = border

                col_con = 2 + len(MESES_ORDEN)
                vcon = subs[sub]["consolidado"]
                cellc = ws.cell(row=row, column=col_con, value=_money_num(vcon) if vcon else "")
                cellc.number_format = '#,##0'
                cellc.alignment = align_right
                cellc.border = border

                if row % 2 == 0:
                    for col in range(1, len(headers) + 1):
                        ws.cell(row=row, column=col).fill = fill_light

                row += 1

            total_cat = {m: 0.0 for m in MESES_ORDEN}
            for sub in subs:
                for m in MESES_ORDEN:
                    total_cat[m] += subs[sub][m]
            total_con = sum(total_cat[m] for m in MESES_ORDEN)

            ws.cell(row=row, column=1, value=f"TOTAL {cat}").font = bold
            ws.cell(row=row, column=1).alignment = align_left

            for idx, m in enumerate(MESES_ORDEN, start=2):
                cell = ws.cell(row=row, column=idx, value=_money_num(total_cat[m]) if total_cat[m] else "")
                cell.number_format = '#,##0'
                cell.font = bold
                cell.alignment = align_right
                cell.border = border

            col_con = 2 + len(MESES_ORDEN)
            cellc = ws.cell(row=row, column=col_con, value=_money_num(total_con) if total_con else "")
            cellc.number_format = '#,##0'
            cellc.font = bold
            cellc.alignment = align_right
            cellc.border = border

            for col in range(1, len(headers) + 1):
                ws.cell(row=row, column=col).fill = fill_gray

            row += 2

    escribir_seccion(T["ingresos"], estructura_ing)
    escribir_seccion(T["egresos"], estructura_egr)

    total_ing = float(df_ing["valor"].sum()) if not df_ing.empty else 0.0
    total_egr = float(df_egr["valor"].sum()) if not df_egr.empty else 0.0
    balance = total_ing - total_egr

    ws.cell(row=row, column=1, value=T["resumen"]).font = title_font
    row += 1

    resumen = [
        (T["consolidado_ing"], total_ing),
        (T["consolidado_egr"], total_egr),
        (T["balance"], balance),
    ]
    for label, val in resumen:
        ws.cell(row=row, column=1, value=label).font = bold
        ws.cell(row=row, column=1).alignment = align_left
        ws.cell(row=row, column=1).border = border

        ws.cell(row=row, column=2, value=_money_num(val)).number_format = '#,##0'
        ws.cell(row=row, column=2).alignment = align_right
        ws.cell(row=row, column=2).border = border
        row += 1

    row += 2

    ws.cell(row=row, column=1, value=T["resumen_madre"]).font = title_font
    row += 1

    encabez = [T["codigo_madre"], T["nombre"], T["ingresos"], T["egresos"], T["balance"]]
    for c, h in enumerate(encabez, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = white_bold
        cell.fill = fill_dark
        cell.border = border
        cell.alignment = align_left if c <= 2 else align_right
    row += 1

    if resumen_madres_df is None or resumen_madres_df.empty:
        ws.cell(row=row, column=1, value=T["sin_movimientos"]).alignment = align_left
        for c in range(1, len(encabez) + 1):
            ws.cell(row=row, column=c).border = border
        row += 2
    else:
        for _, r in resumen_madres_df.iterrows():
            ws.cell(row=row, column=1, value=str(r.get("codigoMadre", ""))).border = border
            ws.cell(row=row, column=2, value=str(r.get("nombre", ""))).border = border

            ws.cell(row=row, column=3, value=_money_num(r.get("ingresos", 0))).number_format = '#,##0'
            ws.cell(row=row, column=4, value=_money_num(r.get("egresos", 0))).number_format = '#,##0'
            ws.cell(row=row, column=5, value=_money_num(r.get("balance", 0))).number_format = '#,##0'
            for c in range(3, 6):
                ws.cell(row=row, column=c).alignment = align_right
                ws.cell(row=row, column=c).border = border
            ws.cell(row=row, column=1).alignment = align_left
            ws.cell(row=row, column=2).alignment = align_left
            row += 1

    ws.column_dimensions["A"].width = 45
    for i in range(2, 2 + len(MESES_ORDEN) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 14
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 18

    # Hoja 2: detalle
    ws_det = wb_out.create_sheet("Mother Detail" if lang == "en" else "Detalle por Madres")

    ws_det["A1"] = T["detalle_madre"]
    ws_det["A1"].font = title_font
    ws_det.merge_cells("A1:G1")
    ws_det["A2"] = f"{T['fecha_corte']}:"
    ws_det["B2"] = corte
    ws_det["A2"].font = bold

    rowd = 4
    det_header = [T["fecha"], T["tipo"], T["medio"], T["categoria"], T["subcategoria"], T["concepto"], T["valor"]]

    if detalle_madres_df is None or detalle_madres_df.empty:
        ws_det.cell(row=rowd, column=1, value=T["sin_datos"]).alignment = align_left
    else:
        for cod, subdf in detalle_madres_df.groupby("codigoMadre"):
            try:
                nombre = str(subdf["nombre"].iloc[0] or "").strip()
            except:
                nombre = ""

            ws_det.cell(row=rowd, column=1, value=f"{cod} - {nombre}" if nombre else f"{cod} -").font = bold
            rowd += 1

            for c, h in enumerate(det_header, start=1):
                cell = ws_det.cell(row=rowd, column=c, value=h)
                cell.font = white_bold
                cell.fill = PatternFill(start_color="455A64", end_color="455A64", fill_type="solid")
                cell.border = border
                cell.alignment = align_left if c <= 6 else align_right
            rowd += 1

            for _, rr in subdf.iterrows():
                f = rr.get("fecha_dt", None)
                f_txt = f.strftime("%d/%m/%Y") if isinstance(f, (datetime, date)) and f else str(rr.get("fecha", ""))

                ws_det.cell(row=rowd, column=1, value=f_txt).border = border
                ws_det.cell(row=rowd, column=2, value=traducir_tipo(rr.get("tipo", ""), lang)).border = border
                ws_det.cell(row=rowd, column=3, value=traducir_medio(rr.get("medio", ""), lang)).border = border
                ws_det.cell(
                    row=rowd,
                    column=4,
                    value=_tr_label(rr.get("categoria", ""), lang, CAT_EN)
                ).border = border

                ws_det.cell(
                    row=rowd,
                    column=5,
                    value=_tr_label(rr.get("subcategoria", ""), lang, CAT_EN)
                ).border = border
                ws_det.cell(row=rowd, column=6, value=str(rr.get("concepto", ""))).border = border

                ws_det.cell(row=rowd, column=7, value=_money_num(rr.get("valor", 0))).number_format = '#,##0'
                ws_det.cell(row=rowd, column=7).alignment = align_right
                ws_det.cell(row=rowd, column=7).border = border

                for c in range(1, 7):
                    ws_det.cell(row=rowd, column=c).alignment = align_left
                rowd += 1

            rowd += 2

    ws_det.column_dimensions["A"].width = 14
    ws_det.column_dimensions["B"].width = 10
    ws_det.column_dimensions["C"].width = 12
    ws_det.column_dimensions["D"].width = 22
    ws_det.column_dimensions["E"].width = 22
    ws_det.column_dimensions["F"].width = 40
    ws_det.column_dimensions["G"].width = 14

    output = io.BytesIO()
    wb_out.save(output)
    output.seek(0)

    filename = f"Informe_Mensual_Consolidado_{lang.upper()}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = make_response(output.read())
    response.headers.set("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers.set("Content-Disposition", f"attachment; filename={filename}")
    return response


# -------------------------
#  API OPCIONES (categorías y subcategorías)
# -------------------------
@main_bp.route('/api/opciones/categorias', methods=['GET'])
def api_get_categorias():
    wb = get_workbook(data_only=True)
    categorias = set()
    subcategorias = set()

    if "Ingresos" in wb.sheetnames:
        for row in wb["Ingresos"].iter_rows(min_row=2, values_only=True):
            if row and row[3]:
                categorias.add(str(row[3]))
            if row and row[4]:
                subcategorias.add(str(row[4]))

    if "Egresos" in wb.sheetnames:
        for row in wb["Egresos"].iter_rows(min_row=2, values_only=True):
            if row and row[3]:
                categorias.add(str(row[3]))
            if row and row[4]:
                subcategorias.add(str(row[4]))

    return jsonify({
        "categorias": sorted(list(categorias)),
        "subcategorias": sorted(list(subcategorias))
    })


# ============================
#  ✅ API: Listado de Madres (SIN PISAR /api/madres)
# ============================
@main_bp.route("/api/madres/lista")
def api_madres_lista():
    try:
        df = pd.read_excel(Config.EXCEL_FILE, sheet_name="Registro Madres")
        df.columns = [c.strip().lower() for c in df.columns]

        col_codigo = next((c for c in df.columns if "codigo" in c and "madre" in c), None)
        col_nombre = next((c for c in df.columns if "nombre" in c), None)

        if not col_codigo or not col_nombre:
            return jsonify([])

        madres = []
        for _, row in df.iterrows():
            codigo = _norm_cod_madre(row.get(col_codigo, ""))
            nombre = str(row.get(col_nombre, "")).strip()
            if codigo and codigo.lower() != "nan":
                madres.append({"codigoMadre": codigo, "nombreCompleto": "" if nombre.lower() == "nan" else nombre})

        madres.sort(key=lambda x: x["nombreCompleto"])
        return jsonify(madres)

    except Exception:
        return jsonify([]), 500


import json

# ============================
#   LOGIN / LOGOUT / SESIÓN
# ============================

@main_bp.route('/')
def index():
    session.clear()
    return render_template("login.html")


@main_bp.route('/login')
def login_page():
    return render_template('login.html')


@main_bp.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    usuario = data.get("usuario", "").strip()
    password = data.get("password", "").strip()

    try:
        with open(Config.USUARIOS_FILE, "r", encoding="utf-8") as f:
            usuarios = json.load(f)
    except Exception:
        return jsonify({"error": "Error leyendo usuarios"}), 500

    for u in usuarios:
        if u["usuario"] == usuario and u["password"] == password:
            session["usuario"] = u["usuario"]
            session["rol"] = u["rol"]
            session["nombre"] = u["nombre"]
            return jsonify({"rol": u["rol"], "nombre": u["nombre"]})
    return jsonify({"error": "credenciales inválidas"}), 401


@main_bp.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('index'))


# ============================
#   GESTIÓN DE USUARIOS (solo admin)
# ============================

@main_bp.route('/api/usuarios', methods=['GET'])
def get_usuarios():
    if session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403
    with open(Config.USUARIOS_FILE, "r", encoding="utf-8") as f:
        return jsonify(json.load(f))


@main_bp.route('/api/usuarios', methods=['POST'])
def add_usuario():
    if session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403

    data = request.json
    codigo = data.get("codigo")
    if codigo != "ADMIN-2025":
        return jsonify({"error": "Código incorrecto"}), 401

    with open(Config.USUARIOS_FILE, "r+", encoding="utf-8") as f:
        usuarios = json.load(f)
        usuarios.append({
            "usuario": data["usuario"],
            "nombre": data["nombre"],
            "password": data["password"],
            "rol": data["rol"]
        })
        f.seek(0)
        json.dump(usuarios, f, indent=2, ensure_ascii=False)
    return jsonify({"message": "Usuario agregado"})


@main_bp.route('/api/usuarios/<usuario>', methods=['DELETE'])
def delete_usuario(usuario):
    if session.get("rol") != "admin":
        return jsonify({"error": "No autorizado"}), 403
    data = request.json or {}
    if data.get("codigo") != "ADMIN-2025":
        return jsonify({"error": "Código incorrecto"}), 401

    with open(Config.USUARIOS_FILE, "r+", encoding="utf-8") as f:
        usuarios = json.load(f)
        usuarios = [u for u in usuarios if u["usuario"] != usuario]
        f.seek(0)
        f.truncate()
        json.dump(usuarios, f, indent=2, ensure_ascii=False)
    return jsonify({"message": "Usuario eliminado"})


