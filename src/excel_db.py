import openpyxl
# pyrefly: ignore [missing-import]
from filelock import FileLock
from datetime import datetime, date
from collections import defaultdict
from .config import Config

LOCK_FILE = str(Config.EXCEL_FILE) + ".lock"
excel_lock = FileLock(LOCK_FILE, timeout=10)

def get_workbook(data_only=False):
    """
    Lee el archivo Excel. No requiere lock para lectura pura,
    pero es recomendable si hay escrituras concurrentes. 
    Para mantener el rendimiento, la lectura sin lock es aceptable, 
    pero la carga de datos compartidos debe manejarse con cuidado.
    """
    return openpyxl.load_workbook(Config.EXCEL_FILE, data_only=data_only)

def save_workbook(wb):
    """
    Guarda el archivo de Excel asegurando que nadie más escriba al mismo tiempo.
    """
    with excel_lock:
        wb.save(Config.EXCEL_FILE)

def parse_date_cell(value):
    if value is None:
        return None
    if isinstance(value, (datetime, date)):
        return value.date() if isinstance(value, datetime) else value
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except Exception:
            continue
    try:
        return datetime.fromisoformat(s).date()
    except Exception:
        return None

def row_to_dict(row_tuple, row_num, sheet_name):
    return {
        "id": row_num,
        "sheet": sheet_name,
        "fecha": str(row_tuple[0]) if len(row_tuple) > 0 and row_tuple[0] else "",
        "medio": (row_tuple[1] or "").lower() if len(row_tuple) > 1 else "",
        "tipo": (str(row_tuple[2]).lower().rstrip("s")
                 if len(row_tuple) > 2 and row_tuple[2]
                 else ("ingreso" if sheet_name == "Ingresos" else "egreso")),
        "categoria": row_tuple[3] if len(row_tuple) > 3 and row_tuple[3] else "",
        "subcategoria": row_tuple[4] if len(row_tuple) > 4 and row_tuple[4] else "",
        "codigoMadre": row_tuple[5] if len(row_tuple) > 5 and row_tuple[5] else "",
        "concepto": row_tuple[6] if len(row_tuple) > 6 and row_tuple[6] else "",
        "valor": float(row_tuple[7]) if len(row_tuple) > 7 and row_tuple[7] not in (None, "") else 0.0,
        "responsable": row_tuple[8] if len(row_tuple) > 8 and row_tuple[8] else "",
    }

def actualizar_reportes(wb):
    ingresos_sheet = wb["Ingresos"]
    egresos_sheet = wb["Egresos"]

    def safe_sum(sheet, col_index):
        total = 0.0
        for row in sheet.iter_rows(min_row=2, values_only=True):
            try:
                val = row[col_index]
                if val is None or val == "":
                    continue
                total += float(val)
            except Exception:
                continue
        return total

    total_ingresos = safe_sum(ingresos_sheet, 7)
    total_egresos = safe_sum(egresos_sheet, 7)
    balance = total_ingresos - total_egresos

    if "Resumen Consolidado" in wb.sheetnames:
        resumen = wb["Resumen Consolidado"]
        if resumen.max_row >= 2:
            resumen.delete_rows(2, resumen.max_row)
        resumen.append(["Total Ingresos", total_ingresos])
        resumen.append(["Total Egresos", total_egresos])
        resumen.append(["Balance", balance])

    ingresos_por_mes = defaultdict(float)
    for row in ingresos_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[7]:
            fecha = parse_date_cell(row[0])
            if fecha:
                clave = fecha.strftime("%Y-%m")
                ingresos_por_mes[clave] += float(row[7] or 0)

    egresos_por_mes = defaultdict(float)
    for row in egresos_sheet.iter_rows(min_row=2, values_only=True):
        if row and row[0] and row[7]:
            fecha = parse_date_cell(row[0])
            if fecha:
                clave = fecha.strftime("%Y-%m")
                egresos_por_mes[clave] += float(row[7] or 0)

    if "Informe Mensual" in wb.sheetnames:
        informe = wb["Informe Mensual"]
        if informe.max_row >= 2:
            informe.delete_rows(2, informe.max_row)
        informe.append(["Mes", "Ingresos", "Egresos", "Balance"])
        meses = sorted(set(ingresos_por_mes.keys()) | set(egresos_por_mes.keys()))
        for mes in meses:
            ingreso = ingresos_por_mes.get(mes, 0)
            egreso = egresos_por_mes.get(mes, 0)
            informe.append([mes, ingreso, egreso, ingreso - egreso])
