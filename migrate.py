import pandas as pd
from openpyxl import load_workbook

file_path = 'data/adonai_data_completo.xlsx'
wb = load_workbook(file_path)

# --- Mapeo de INGRESOS ---
def map_ingresos(row):
    # column index: Medio=2, Tipo=3, Categoria=4, Subcat=5, Concepto=7  (1-based in openpyxl)
    # openpyxl uses 1-based indexing
    # But wait, in the codebase:
    # 1=Fecha, 2=Medio, 3=Tipo, 4=Categoria, 5=Subcategoria, 6=CodigoMadre, 7=Concepto, 8=Valor, 9=Responsable
    medio = (row[1].value or "").strip()
    cat = (row[3].value or "").strip()
    subcat = (row[4].value or "").strip()
    
    nuevo_medio = medio
    nueva_cat = cat
    nueva_subcat = subcat
    
    if subcat == "Intereses banco":
        nueva_cat = "Administrativo"
        nueva_subcat = "Intereses bancarios"
    elif subcat == "Sostenimiento SRL":
        nueva_cat = "Administrativo"
        nueva_subcat = "Donaciones SRL"
    elif subcat == "Donaciones locales":
        nueva_cat = "Misión Institucional"
        nueva_subcat = "Donaciones locales"
    elif subcat == "Donaciones específicas":
        nueva_cat = "Misión Institucional"
        nueva_subcat = "Donaciones específicas locales"
    elif subcat == "Donación USA":
        nueva_cat = "Misión Institucional"
        nueva_subcat = "Donaciones USA"
    elif subcat == "Boutique ventas efectivo":
        nueva_cat = "Unidad de Negocio"
        nueva_subcat = "Donaciones ropa"
        nuevo_medio = "caja"
    elif subcat == "Boutique ventas banco":
        nueva_cat = "Unidad de Negocio"
        nueva_subcat = "Donaciones ropa"
        nuevo_medio = "banco"
    elif subcat in ["Capacitaciones", "Taller", "Comercialización"]:
        nueva_cat = "Unidad de Negocio"
        nueva_subcat = subcat
    elif subcat == "Otros ingresos" or cat == "Otros ingresos":
        nueva_cat = "Otros ingresos"
        nueva_subcat = "Otros ingresos"
    elif cat == "Ministerial":
        nueva_cat = "Otros ingresos"
        nueva_subcat = "Otros ingresos"
        
    return nuevo_medio, nueva_cat, nueva_subcat

# --- Mapeo de EGRESOS ---
def map_egresos(row):
    cat = (row[3].value or "").strip()
    subcat = (row[4].value or "").strip()
    
    nueva_cat = cat
    nueva_subcat = subcat
    
    if cat == "Administrativo - Nómina":
        nueva_cat = "1. Nómina"
    elif cat == "Administrativo - Servicios":
        if subcat == "Colaboradoras":
            nueva_cat = "2. Colaboradoras Internas"
        else:
            nueva_cat = "3. Proveedores"
    elif cat == "Administrativo - Operacionales":
        nueva_cat = "4. Operacionales"
    elif cat == "Administrativo - Redes Sociales":
        nueva_cat = "5. Redes Sociales"
    elif cat == "Administrativo - Material publicitario":
        nueva_cat = "6. Material Publicitario"
    elif cat == "Administrativo - Trabajo de campo":
        nueva_cat = "7. Trabajo de Campo"
    elif cat == "Administrativo - Otros gastos":
        nueva_cat = "8. Otros Gastos"
    elif cat == "Misión institucional - Bebés":
        nueva_cat = "9. Misional Bebés"
    elif cat == "Misión institucional - Madres":
        nueva_cat = "10. Misional Madres"
    elif cat == "Otros":
        nueva_cat = "11. Otros"
    elif cat == "Unidad de negocio":
        nueva_cat = "12. Unidad de Negocio"
        
    return nueva_cat, nueva_subcat

# Procesar Ingresos
if "Ingresos" in wb.sheetnames:
    ws_ing = wb["Ingresos"]
    for row in ws_ing.iter_rows(min_row=2):
        if not row[0].value: continue
        nuevo_medio, nueva_cat, nueva_subcat = map_ingresos(row)
        row[1].value = nuevo_medio
        row[3].value = nueva_cat
        row[4].value = nueva_subcat

# Procesar Egresos
if "Egresos" in wb.sheetnames:
    ws_eg = wb["Egresos"]
    for row in ws_eg.iter_rows(min_row=2):
        if not row[0].value: continue
        nueva_cat, nueva_subcat = map_egresos(row)
        row[3].value = nueva_cat
        row[4].value = nueva_subcat

wb.save(file_path)
print("Migración completada con éxito.")
